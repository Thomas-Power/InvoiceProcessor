import shutil
import shutil, os
import openpyxl
global holiday
global wbook
global wbname
global currentMonth
global f
global loopcheck
weekday = {"Mon": "y0", "Tue": "y1", "Wed": "y2", "Thu": "y3", "Fri": "n4"}
dayweek = ["Mon", "Tue", "Wed", "Thu", "Fri"]
holidays = openpyxl.load_workbook('.\\Holidays.xlsx')
hsheet = holidays.get_sheet_by_name('Sheet1')


for wbname in os.listdir('.\\Output2'):
	wbook = openpyxl.load_workbook('.\\Output2\\' + str(wbname))
	currentMonth = wbook.get_sheet_by_name('0617')
	#reformat dates
	for i in range(14, 50, 1):
		datechng = str(currentMonth.cell(row=i, column=1).value)
		if len(datechng) == 10:
			currentMonth.cell(row=i, column=1).value = datechng[8:10] + "-" + "Dec-17"
		if len(datechng) == 9:
			currentMonth.cell(row=i, column=1).value = "0" + datechng[8] + "-" + "Dec-17"
			
	for i in range(14, 50, 1):
		day = currentMonth.cell(row=i, column=1).value
		daystr = str(currentMonth.cell(row=i, column=2).value)[:3]
		if currentMonth.cell(row=i, column=2).value != None:
			if currentMonth.cell(row=i, column=5).value != None:
				if currentMonth.cell(row=i, column=1).value != None:
					if "30-Oct-17" in str(day):
						#Current day is not a friday
						if "y" == weekday[daystr][0]:
							if currentMonth.cell(row=(i + 1), column=5).value != None:
								#Following date is next day
								if int(weekday[daystr][1]) + 1 == int(weekday[str(currentMonth.cell(row=(i + 1), column=2).value)[:3]][1]):
									#Following day is worth more than current day
									if float(str(currentMonth.cell(row=i, column=5).value)) <= float(str(currentMonth.cell(row=(i + 1), column=5).value)):
										#Add +1 day to all dates where no. day + 1 > day + 2
										currentMonth.cell(row=i, column=1).value = str(currentMonth.cell(row=(i + 1), column=1).value)
										currentMonth.cell(row=i, column=2).value = str(currentMonth.cell(row=(i + 1), column=2).value)
										currentMonth.cell(row=i, column=5).value = currentMonth.cell(row=(i + 1), column=5).value
										loopcheck = 0
										for e in range(i + 1, 45, 1):
											if loopcheck == 0:
												if currentMonth.cell(row=(e + 1), column=2).value != None:
													if int(weekday[str(currentMonth.cell(row=(e), column=2).value)[:3]][1:]) < int(weekday[str(currentMonth.cell(row=(e + 1), column=2).value)[:3]][1:]):
														if int(str(currentMonth.cell(row=(e), column=1).value)[1]) + 1 == 10:
															if int(str(currentMonth.cell(row=(e), column=1).value)[0]) == 2:
																currentMonth.cell(row=(e), column=1).value = "30" + str(currentMonth.cell(row=(e), column=1).value)[2:]
															if int(str(currentMonth.cell(row=(e), column=1).value)[0]) == 1:
																currentMonth.cell(row=(e), column=1).value = "20" + str(currentMonth.cell(row=(e), column=1).value)[2:]
															if str(currentMonth.cell(row=(e), column=1).value)[0] == 0:
																currentMonth.cell(row=(e), column=1).value = "10" + str(currentMonth.cell(row=(e), column=1).value)[2:]
														if int(str(currentMonth.cell(row=(e), column=1).value)[:2]) == 31:
															currentMonth.cell(row=(e), column=1).value = ""
															currentMonth.cell(row=(e), column=2).value = ""
															currentMonth.cell(row=(e), column=5).value = ""
														else:
															cday = int(weekday[str(currentMonth.cell(row=e, column=2).value)[:3]][1]) + 1
															if cday != 5:
																currentMonth.cell(row=(e), column=1).value = str(int(str(currentMonth.cell(row=(e), column=1).value)[0])) + str(int(str(currentMonth.cell(row=(e), column=1).value)[1]) + 1) + str(currentMonth.cell(row=(e), column=1).value)[2:]
																currentMonth.cell(row=e, column=2).value = dayweek[cday] + str(currentMonth.cell(row=(e), column=2).value)[3:]
															else:
																pass
													else:
														currentMonth.cell(row=e, column=1).value = currentMonth.cell(row=e + 1, column=1).value
														currentMonth.cell(row=e, column=2).value = currentMonth.cell(row=e + 1, column=2).value
														currentMonth.cell(row=e, column=5).value = currentMonth.cell(row=e + 1, column=5).value
														currentMonth.cell(row=e + 1, column=1).value = ""
														currentMonth.cell(row=e + 1, column=2).value = ""
														currentMonth.cell(row=e + 1, column=5).value = ""
														loopcheck = 1
												
												else:
													currentMonth.cell(row=e, column=1).value = currentMonth.cell(row=(e + 1), column=1).value
													currentMonth.cell(row=e, column=2).value = currentMonth.cell(row=(e + 1), column=2).value
													currentMonth.cell(row=e, column=5).value = currentMonth.cell(row=(e + 1), column=5).value
												
										
									#Following date is next day, value does not exceed current day, copy current day over following day
									else:
										currentMonth.cell(row=i, column=1).value = str(currentMonth.cell(row=(i + 1), column=1).value)
										currentMonth.cell(row=i, column=2).value = str(currentMonth.cell(row=(i + 1), column=2).value)[:3] + str(currentMonth.cell(row=i, column=2).value)[3:]
										#Add +1 day to all dates where no. day + 1 > day + 2
										
										loopcheck = 0
										for e in range(i + 2, 45, 1):
											if loopcheck == 0:
												if currentMonth.cell(row=(e + 1), column=2).value != None:
													if int(weekday[str(currentMonth.cell(row=(e), column=2).value)[:3]][1:]) < int(weekday[str(currentMonth.cell(row=(e + 1), column=2).value)[:3]][1:]):
														
														if int(str(currentMonth.cell(row=(e), column=1).value)[1]) + 1 == 10:
															if int(str(currentMonth.cell(row=(e), column=1).value)[0]) == 2:
																currentMonth.cell(row=(e), column=1).value = "30" + str(currentMonth.cell(row=(e), column=1).value)[2:]
															if int(str(currentMonth.cell(row=(e), column=1).value)[0]) == 1:
																currentMonth.cell(row=(e), column=1).value = "20" + str(currentMonth.cell(row=(e), column=1).value)[2:]
															if str(currentMonth.cell(row=(e), column=1).value)[0] == 0:
																currentMonth.cell(row=(e), column=1).value = "10" + str(currentMonth.cell(row=(e), column=1).value)[2:]
														if int(str(currentMonth.cell(row=(e), column=1).value)[:2]) == 31:
															currentMonth.cell(row=(e), column=1).value = ""
															currentMonth.cell(row=(e), column=2).value = ""
															currentMonth.cell(row=(e), column=5).value = ""
														else:
															cday = int(weekday[str(currentMonth.cell(row=e, column=2).value)[:3]][1]) + 1
															if cday != 5:
																if int(weekday[str(currentMonth.cell(row=(e), column=2).value)[:3]][1]) != int((weekday[str(currentMonth.cell(row=(e + 1), column=2).value)[:3]][1])):
																	currentMonth.cell(row=(e), column=1).value = str(int(str(currentMonth.cell(row=(e), column=1).value)[0])) + str(int(str(currentMonth.cell(row=(e), column=1).value)[1]) + 1) + str(currentMonth.cell(row=(e), column=1).value)[2:]
																	currentMonth.cell(row=e, column=2).value = dayweek[cday] + str(currentMonth.cell(row=(e), column=2).value)[3:]
															else: 
																pass
															
													else:
														currentMonth.cell(row=e, column=1).value = currentMonth.cell(row=e + 1, column=1).value
														currentMonth.cell(row=e, column=2).value = currentMonth.cell(row=e + 1, column=2).value
														currentMonth.cell(row=e, column=5).value = currentMonth.cell(row=e + 1, column=5).value
														currentMonth.cell(row=e + 1, column=1).value = ""
														currentMonth.cell(row=e + 1, column=2).value = ""
														currentMonth.cell(row=e + 1, column=5).value = ""
														loopcheck = 1
															
													
												else:
													currentMonth.cell(row=e, column=1).value = currentMonth.cell(row=(e + 1), column=1).value
													currentMonth.cell(row=e, column=2).value = currentMonth.cell(row=(e + 1), column=2).value
													currentMonth.cell(row=e, column=5).value = currentMonth.cell(row=(e + 1), column=5).value
													
								#Following date is not next day
							else:
									#Add +1 day to all dates where no. day >= day + 1
								currentMonth.cell(row=(i), column=1).value = str(int(str(currentMonth.cell(row=(i), column=1).value)[0])) + str(int(str(currentMonth.cell(row=(i), column=1).value)[1]) + 1) + str(currentMonth.cell(row=(i), column=1).value)[2:]
								currentMonth.cell(row=(i), column=2).value = dayweek[int(weekday[str(currentMonth.cell(row=i, column=2).value)[:3]][1]) + 1] + str(currentMonth.cell(row=(i), column=2).value)[3:]
									
									
									
	originalname = os.path.basename('.\\Output2\\' +str(wbname))
	wbook.save('.\\Output2\\' + originalname)
	shutil.move('.\\Output2\\' + str(wbname), '.\\Output3\\' + originalname)