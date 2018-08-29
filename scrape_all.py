#Copy Template onto file in output folder
import sys
import shutil
import shutil, os
import openpyxl
import pymongo
from pymongo import MongoClient
import datetime 
import calendar

client = MongoClient("localhost", 27017)
db = client.Bluebell_Dairies
customers = db.customers
orders = db.orders

#If June sheet available 
days = ["monday", "tuesday", "wednesday", "thursday", "friday"]

def customer_data():
	return {"name": "",
				"customer_no": None,
				"address": ""
	}
	
def order_data():
	return {
				"month": None,
				"year": None,
				"customer_no": None,
				"reference_no": None,
				"delivery_charge": None,
				"monday": {
					"descriptions": "",
					"value": None
				}, 
				"tuesday": {
					"descriptions": "",
					"value": None
				}, 
				"wednesday": {
					"descriptions": "",
					"value": None
				}, 
				"thursday": {
					"descriptions": "",
					"value": None
				}, 
				"friday": {
					"descriptions": "",
					"value": None
				}
	}

def thirdWeek(year, month):
	result = {}
	c = calendar.Calendar(firstweekday=calendar.SUNDAY)
	monthcal = c.monthdatescalendar(year,month)
	third_friday = [day for week in monthcal for day in week if \
					day.weekday() == calendar.FRIDAY and \
					day.month == month][2]
	for i in range(5):
		cur_day = str(third_friday)[:-2] + str(int(str(third_friday)[-2:]) - i)
		result.update({str(datetime.datetime.strptime(cur_day, '%Y-%m-%d').strftime('%d-%b-%y')):4-i})
		result.update({str(datetime.datetime.strptime(cur_day, '%Y-%m-%d').strftime('%d-%m-%y')):4-i})
		result.update({str(datetime.datetime.strptime(cur_day, '%Y-%m-%d').strftime('%d/%b/%y')):4-i})
		result.update({str(datetime.datetime.strptime(cur_day, '%Y-%m-%d').strftime('%d/%m/%y')):4-i})
		result.update({str(datetime.datetime.strptime(cur_day, '%Y-%m-%d').strftime('%Y-%m-%d')):4-i})
		result.update({str(datetime.datetime.strptime(cur_day, '%Y-%m-%d').strftime('%Y/%m/%d')):4-i})
	return result

		
if __name__ == '__main__':
	print("[PYTHON]: Starting process...")
	sys.stdout.flush()
	cur_year = int(sys.argv[1])
	cur_month = int(sys.argv[2])

	date = thirdWeek(cur_year, cur_month)
	cur_year = int(str(cur_year)[-2:])
	file_dir = '.\\Invoices\\' + str(cur_month) + "-" + str(cur_year) + "\\"
	for wbname in os.listdir(file_dir):
		cur_file = customer_data()
		order = order_data()
		wbook = openpyxl.load_workbook(file_dir + str(wbname))
		try:
			sheet = wbook['0617']
		except KeyError:
			sheet = wbook['Sheet1']
		customer_no = int(wbname[:3])
		cur_file["name"] = str(sheet.cell(row=6, column=2).value)
		cur_file["customer_no"] = customer_no
		try:
			June = wbook.get_sheet_by_name('0617')
			#Copy addresses from wb B6-B12 onto '.\\file.xlsx' B6-B12
			check = False
			for i in range(8, 13, 1):
				if(len(str(June.cell(row=i, column=2).value)) > 1):
					if(check):
						cur_file["address"] = cur_file["address"] + "," + str(June.cell(row=i, column=2).value)
					else:
						cur_file["address"] = str(June.cell(row=i, column=2).value)
						check = True
		except KeyError:
			try:
				for i in range(6, 13, 1):
					cur_file["address"] = cur_file["address"] + "," + str(sheet.cell(row=i, column=2).value)
			except KeyError:
				pass
		if(customers.find_one({"customer_no": customer_no}) == None):
			customers.insert_one(cur_file)
		else:
			customers.update_one({"customer_no": customer_no}, {"$set": cur_file}, upsert=False)
		for i in range(14, 50, 1):
			daycheck = (str(sheet.cell(row=i, column=1).value).split(" "))[0]
			try:		
				e = date[str(daycheck)]
				descriptions = str(sheet.cell(row=(i), column=2).value)
				money = sheet.cell(row=(i), column=5).value
				order[days[e]]["descriptions"] = descriptions[4:]
				order[days[e]]["value"] = money
			except ValueError:
				pass
			except KeyError:
				pass
		for aref in range(2, 5, 1):
			if str(sheet.cell(row=aref, column=4).value) == str("Reference no."):
				order["reference_no"] = int(sheet.cell(row=aref, column=5).value)
		if not(sheet.cell(row=50, column=5).value == None):
			try:
				order["delivery_charge"] = int(sheet.cell(row=50, column=5).value)
			except ValueError:
				pass
		order["customer_no"] = customer_no
		order["month"] = cur_month
		order["year"] = cur_year
		if(orders.find_one({'customer_no':customer_no, 'month':cur_month, 'year':cur_year}) == None):
			orders.insert_one(order)
		else:
			orders.update_one({'customer_no':customer_no, 'month':cur_month, 'year':cur_year}, {"$set": order}, upsert=False)
	print("[PYTHON]: Process complete, invoice data scraped for " + str(cur_month) + "-" + str(cur_year))
	sys.stdout.flush()