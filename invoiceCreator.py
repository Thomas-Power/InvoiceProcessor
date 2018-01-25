import shutil
import shutil, os
import openpyxl
global wbname
global currentMonth
global datasheet
global f
global e
global amount
global gap

for wbname in os.listdir('.\\Output'):
	wbook = openpyxl.load_workbook('.\\Output\\' + str(wbname))
	datasheet = wbook.get_sheet_by_name('Sheet1')
	currentMonth = wbook.get_sheet_by_name('0617')
	#Add day strings onto template
	for i in range(14, 19, 1):
		if datasheet.cell(row=i, column=2).value != None:
			datasheet.cell(row=i, column=2).value = str(datasheet.cell(row=i, column=2).value)
			#Copy Delivery charge from E50 April to E22"""
	#Create invoice number (April + (April - March))
	try:
		currentMonthnumber = (int(datasheet.cell(row=4, column=5).value) + int(datasheet.cell(row=4, column=5).value) - int(datasheet.cell(row=5, column=5).value))
		currentMonth.cell(row=2, column=5).value = currentMonthnumber
	except TypeError:
		currentMonth.cell(row=2, column=5).value = ""
		currentMonth.cell(row=2, column=4).value = ""
		pass
	#Copy Del charge over
	currentMonth.cell(row=50, column=5).value = datasheet.cell(row=22, column=5).value
		#If del charge = "" delete D50
	if currentMonth.cell(row=50, column=5).value == None:
		currentMonth.cell(row=50, column=4).value = ""
	#Copy address over
	for i in range(6, 13, 1):
		currentMonth.cell(row=i, column=2).value = datasheet.cell(row=i, column=2).value
	
	#If space available print balance for day
	##
	#e= start day list value
	e = 4
	gap = 0
	for i in range(14, 50, 1):
		if datasheet.cell(row=(14 + e), column=2).value != None or "":
			currentMonth.cell(row=(i - gap), column=2).value = datasheet.cell(row=(14 + e), column=2).value
			currentMonth.cell(row=(i - gap), column=5).value = datasheet.cell(row=(14 + e), column=5).value
			e = e + 1
			if e == 5:
				e = 0
		else:
			gap = gap + 1
			e = e + 1
			if e == 5:
				e = 0
	#Calculate date based on anchor date through associated weekday
	f = 0
	
	for i in range(14, 50, 1):
		day = ["Fri", "Sat", "Sun", "Mon", "Tue", "Wed", "Thu"]
		basedate = "2017-12-01" 
		if currentMonth.cell(row=i, column=2).value != None or "":
			currentMonth.cell(row=i, column=1).value = str(basedate[:8] + str(int(basedate[8:]) + (f + day.index(str(str(currentMonth.cell(row=i, column=2).value)[:3])))))
			
				#if date exceeds work week procedurally move to next
			if currentMonth.cell(row=(i-1), column=1).value != None or "":
				if (i-1) != 13:
				#print(str(int(str(currentMonth.cell(row=(i-1), column=1).value)[8:]))
					x = int(str(currentMonth.cell(row=i, column=1).value)[8:])
					y = int(str(currentMonth.cell(row=(i-1), column=1).value)[8:])
					if x <= y:
						f = f + 7
						date = basedate[:8] + str(int(basedate[8:]) + (f + day.index(str(currentMonth.cell(row=i, column=2).value)[:3])))
						currentMonth.cell(row=i, column=1).value = date
						
						# Delete any dates over month of currentMonth, clear all proceeding data, end loop
		try:
			if int(str(currentMonth.cell(row=i, column=1).value)[8:]) >= 23:
				currentMonth.cell(row=i, column=1).value = ""
				currentMonth.cell(row=i, column=2).value = ""
				currentMonth.cell(row=i, column=5).value = ""
				for g in range (i, 50, 1):
					currentMonth.cell(row=g, column=1).value = ""
					currentMonth.cell(row=g, column=2).value = ""
					currentMonth.cell(row=g, column=5).value = ""
				break
		except ValueError:
			break

	#calculate total
	amount = 0
	wbook.remove_sheet(datasheet)
	"""for i in range(14, 51, 1):
		if currentMonth.cell(row=i, column=5).value != None or "":
			try:
				amount = amount + float(str(currentMonth.cell(row=i, column=5).value))
			except ValueError:
				try:
					amount = amount + int(str(currentMonth.cell(row=i, column=5).value))
				except ValueError:
					pass"""
					
	#currentMonth.cell(row=51, column=5).value = amount
	
	originalname = os.path.basename('.\\Output\\' +str(wbname))
	wbook.save('.\\Output\\' + originalname)
	shutil.move('.\\Output\\' + str(wbname), '.\\Output2\\' + originalname)