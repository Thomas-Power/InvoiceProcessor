#Copy Template onto file in output folder
import shutil
import shutil, os
import openpyxl
import pymongo
from pymongo import MongoClient
import datetime 
import calendar
from datetime import timedelta
	
now = datetime.datetime.now()
cur_month = now.strftime("%m")
cur_year = now.strftime("%y")
prev_month = (now.replace(day=1) - timedelta(days=1)).strftime("%m")
prev_year = (now.replace(day=1) - timedelta(days=1)).strftime("%y")
prev_prev_month = ((now.replace(day=1) - timedelta(days=1)).replace(day=1) - timedelta(days=1)).strftime("%m")
prev_prev_year = ((now.replace(day=1) - timedelta(days=1)).replace(day=1) - timedelta(days=1)).strftime("%y")

client = MongoClient("localhost", 27017)
db = client.Bluebell_Dairies
customers = db.customers
orders = db.orders
holidays = db.holidays.distinct("date")

def invoice_order(date, order, value):
	return  {
				"date":date,
				"descriptions":order,
				"value":value
			}

for customer in customers.find():
	wbook = openpyxl.load_workbook('.\\invoice_template.xlsx')
	sheet = wbook['Sheet1']
	order = orders.find_one({'customer_no':customer["customer_no"], 'month':int(prev_month), 'year':int(prev_year)})
	
	invoice_orders = []
	insert = 0
	skip_day = 0
	days_month = calendar.monthrange(int("20" + cur_year),int(cur_month))[1]
	
	for days in range(1, days_month + 1):
		day = days + skip_day
		cur_day = datetime.datetime(int("20" + cur_year), int(cur_month), day)
		weekday = cur_day.strftime("%A").lower()
		if(weekday == "saturday"):
			skip_day = 0
		try:
			if((weekday not in ["saturday", "sunday"]) and order[weekday]["descriptions"] != "" and len(order[weekday]["descriptions"]) > 5):
				if cur_day.strftime('%d-%b-%y') in holidays:
					prev_day = cur_day - datetime.timedelta(1)
					prev_weekday = prev_day.strftime("%A").lower()
					if((prev_weekday not in ["saturday", "sunday"]) and 
					(order[weekday]["value"] > order[prev_weekday]["value"]) and
					insert > 0): #replace current days value with previous day
						if(order[weekday]["value"] > order[prev_weekday]["value"]):
							invoice_orders[insert-1] = invoice_order(
								cur_day.strftime('%d-%b-%y'), 
								order[weekday]["descriptions"], 
								order[weekday]["value"]
								)
					else:
						next_day = cur_day + datetime.timedelta(1)
						next_weekday = next_day.strftime("%A").lower()
						if((prev_weekday not in ["saturday", "sunday"]) and
						(order[weekday]["value"] > order[next_weekday]["value"]) and
						day < days_month): #replace current days value with next days
							invoice_orders.append(invoice_order(
								next_day.strftime('%d-%b-%y'),
								order[weekday]["descriptions"],
								order[weekday]["value"]
							))
							skip_day += 1
							insert += 1
							
				else: #regular procedure here
					invoice_orders.append(invoice_order(
						cur_day.strftime('%d-%b-%y'),
						order[weekday]["descriptions"],
						order[weekday]["value"]
					))
					insert += 1
		
		except KeyError:
			pass
	
	sheet.cell(row=6, column=2).value = customer["name"]
	sheet.cell(row=7, column=2).value = "Customer no. " + str(customer["customer_no"])
	
	address = customer["address"].split(",") 
	addr_row = 8
	for line in address:
		if not line == None and not line == "None" and addr_row < 13:
			sheet.cell(row=addr_row, column=2).value = line
			addr_row += 1
		
	if not(order["reference_no"] == None):
		prev_order = orders.find_one({'customer_no':customer["customer_no"], 'month':int(prev_prev_month), 'year':int(prev_prev_year)})
		sheet.cell(row=4, column=4).value = "Reference no:"
		sheet.cell(row=4, column=5).value = order["reference_no"] + (order["reference_no"] - prev_order["reference_no"])

	cur_row = 14
	for cur_order in invoice_orders:
		sheet.cell(row=cur_row, column=1).value = cur_order["date"]
		sheet.cell(row=cur_row, column=2).value = cur_order["descriptions"]
		sheet.cell(row=cur_row, column=5).value = cur_order["value"]
		cur_row += 1
	
	if not(order["delivery_charge"] == None):
		sheet.cell(row=50, column=5).value = order["delivery_charge"]
	else:
		sheet.cell(row=50, column=4).value = ""

	wbook.save('.\\Output\\' + str(customer["customer_no"]) + "- " + customer["name"] + ".xlsx")
