#Copy Template onto file in output folder
import shutil
import shutil, os
import openpyxl
shutil.copy('template.xlsx', '.\\file.xlsx')

#If currentMonth sheet available 
global file
global sheet
file = openpyxl.load_workbook('.\\file.xlsx')
sheet = file.get_sheet_by_name('Sheet1')
date = ['20-Nov-17', '21-Nov-17', '22-Nov-17', '23-Nov-17', '24-Nov-17']
global wbname

for wbname in os.listdir('.\\Input'):
	wbook = openpyxl.load_workbook('.\\Input\\' + str(wbname))
	try:
		global previousMonth
		previousMonth = wbook.get_sheet_by_name('0617')

	try:
		currentMonth = wbook.get_sheet_by_name('0617')
		#Copy addresses from wb B6-B12 onto '.\\file.xlsx' B6-B12
		for i in range(6, 13, 1):
			sheet.cell(row=i, column=2).value = currentMonth.cell(row=i, column=2).value
	except KeyError:
		try:
			for i in range(6, 13, 1):
				sheet.cell(row=i, column=2).value = previousMonth.cell(row=i, column=2).value
		except KeyError:
			pass
	#If previousMonth 17 exists
	for wipe in range(14, 19, 1):
		sheet.cell(row=(wipe), column=2).value = ""
	for wipe2 in range(2, 6, 1):
		sheet.cell(row=(wipe2), column=5).value = ""
	for wipe3 in range(14, 19, 1):
		sheet.cell(row=(wipe3), column=5).value = ""
	for i in range(14, 50, 1): # correct
		daycheck = previousMonth.cell(row=i, column=1).value
		try:		
			e = date.index(str(daycheck))
			descriptions = previousMonth.cell(row=(i), column=2).value
			money = previousMonth.cell(row=(i), column=5).value
			sheet.cell(row=(14 + e), column=2).value = descriptions
			sheet.cell(row=(14 + e), column=5).value = money
		except ValueError:
			pass
	
	sheet.cell(row=22, column=5).value = previousMonth.cell(row=50, column=5).value
	#Copy Invoice Number for previousMonth 'E4' to 'E4'
	for aref in range(2, 5, 1):
		if str(previousMonth.cell(row=aref, column=4).value) == str("Reference no."):
			sheet.cell(row=4, column=5).value = previousMonth.cell(row=aref, column=5).value
	#'E4' previousMonth - 'E4' previousMonth = Invoice jump/Month

	for mref in range(2, 5, 1):
		try:
			March = wbook.get_sheet_by_name('0617')
			if str(March.cell(row=mref, column=4).value) == "Reference no.":
				sheet.cell(row=5, column=5).value = March.cell(row=mref, column=5).value
		except KeyError:
			pass

	file.save('.\\file.xlsx')
	#Get Title in Input Folder
	originalname = os.path.basename('.\\Input\\' +str(wbname))
	shutil.move('.\\file.xlsx', '.\\Output\\' + originalname)
	#Done