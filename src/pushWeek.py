'''
Created on 26 Mar 2018

@author: Thomas
'''
from src import *
import openpyxl

def pushWeek():
    '''scan order book find first space'''
    for i in range(5, 0, -1):
        if (data.orderSheet.cell(row=2, column=i).value != None):
            
            '''if space exists after current day'''
            if(i > makeInvoice.curDay):
                makeInvoice.cov += 1
                makeInvoice.gap += 1
                break
            
            '''if space exists before current day'''
            if(i < makeInvoice.curDay):
                swapValues(i - makeInvoice.curDay)
                break
            
            '''if no space exists'''
            if(i == 1):
                '''check if next day is worth less'''
                if(data.orderSheet.cell(row=5, column=i).value > data.orderSheet.cell(row=5, column=(i+1)).value):
                    swapValues(0)
                    makeInvoice.cov  += 1
                
                '''check if previous day is worth less'''
                if(data.orderSheet.cell(row=5, column=i).value > data.orderSheet.cell(row=5, column=(i-1)).value):
                    swapValues(-1)
                    makeInvoice.gap  += 1
                
                else:
                    makeInvoice.gap  += 1