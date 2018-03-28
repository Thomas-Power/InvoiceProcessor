'''
Created on 26 Mar 2018

@author: Thomas
'''
from src import *
import openpyxl

def postData():
    '''post date'''
    data.invoiceSheet.cell(row=makeInvoice.curRow, column=1).value = formatDate(makeInvoice.curDate)
    '''post order'''
    data.invoiceSheet.cell(row=makeInvoice.curRow, column=2).value = data.orderSheet.cell(row=makeInvoice.curDay, column=2).value
    '''post value'''
    data.invoiceSheet.cell(row=makeInvoice.curRow, column=5).value = data.orderSheet.cell(row=makeInvoice.curDay, column=5).value